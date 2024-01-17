'''
Author       : xianjie.xia
LastEditors  : xianjie.xia
Date         : 2023-08-16 19:02
LastEditTime : 2023-08-21 15:26
description  : 
'''
import json
import logging
import os
import time

import chardet
import openpyxl

DiscardedGuidArray = []

DiscardedGuidMap = {
    # '104': 'Sound',
    # '109': 'PlayerStart',
    # '110': 'SkyLight',
    # '111': 'Sky_Sphere',
    # '112': 'Anchor',
    # '113': 'Trigger',
    # '115': 'HotWeapon',
    # '116': 'Interactor',
    # '117': 'BlockingArea',
    # '118': 'Rotator',
    # '119': 'Skill',
    # '120': 'SphereTrigger',
    # '2695': 'PostProcess',
    # '2696': 'Equipment',
    # '2959': 'CameraTarget',
    # '4118': 'SmoothSync',
    # '4301': 'PointLight',
    # '4302': 'RectLight',
    # '4303': 'SpotLight',
    # '4304': 'PhysicsCable',
    # '4305': 'PhysicsStick',
    # '4306': 'RelativeEffect',
    # '8444': 'DirectionalLight',
    '12683': 'SwimmingVolume',
    '14090': '运动器/冲量对象',
    '14197': 'EffectLogical',
    '14971': 'SkyBox',
    '15848': 'PhysicsRotator',
    '16032': 'PhysicsSpring',
    '16037': 'UIWidget',
    '16038': 'WheeledVehicle4W',
    '20006': 'PhysicsAngularMotor',
    '20191': 'Thruster',
    '20192': 'PhysicsLinearMotor',
    '20193': 'NavMeshVolume',
    '20194': 'NavModifierVolume',
    '20504': 'PhysicsFulcrum',
    '20638': 'HotWeapon',
    '21045': 'PhysicsPrism',
    '21046': 'PhysicsCylinder',
    '21060': 'PhysicsRotation',
    '21151': 'PostProcessAdvance',
    '21378': 'AbilityObject',
    '25782': 'Anchor',
    '28449': 'HumanoidObject',
    '30829': 'VehicleCamera',
    '31479': 'VehicleCameraSetting',
    '31969': 'NPC',
    '47850': 'InteractiveObject',
    '67455': 'PhysicsImpulse',
    '108547': 'ProjectileLauncher',
    '119918': 'PhysicsSports',
    '124744': 'NPC',
}

floders = ["Materials", "Prefabs", "Levels",
           "TempPrefabs", "JavaScripts", "UI", 'Character']
XlsxFilePath = '废弃表.xlsx'
resXlsx = '替换表.xlsx'
DelMat = False
files = ['.ts', '.mat', '.level', '.prefab', '.ui', '.asset']

all_path = []
result_log = []
result_warning = []
result_error = []
mat_error = {}
mats = {}
guids = []
useMats = []
tslist = []
replaceMap = {}


def xlsx_init():
    if not os.path.exists(XlsxFilePath):
        return False
    wb = openpyxl.load_workbook(XlsxFilePath, data_only=True)
    sheetNames = wb.sheetnames
    for sheetname in sheetNames:
        sheet = wb[sheetname]
        rowCount = sheet.max_row
        for row in range(2, rowCount + 1):
            cellValue = sheet.cell(row, 1).value
            if cellValue == None:
                continue
            value = str(cellValue)
            if value in DiscardedGuidArray:
                continue
            DiscardedGuidArray.append(value)
            # continue
    if not os.path.exists(resXlsx):
        return False
    wb = openpyxl.load_workbook(resXlsx, data_only=True)
    sheetName = wb.sheetnames[0]
    sheet = wb[sheetName]
    rowCount = sheet.max_row
    keys = ['oid', 'type', 'rid', 'name', 'path', 'mesh', 'mat']
    for row in range(2, rowCount + 1):
        key = sheet.cell(row, 1).value
        if key == None:
            continue
        line = {}
        col = 1
        for k in keys:
            v = sheet.cell(row, col).value
            if v and v != '0':
                line[k] = str(v)
            elif k == 'type':
                col = 0
                break
            col += 1
        if col > 0:
            replaceMap[str(key)] = line
    return True


def get_file_encoding(in_path):
    f = open(in_path, 'rb')
    data = f.read()
    file_encoding = chardet.detect(data).get('encoding')
    f.close()
    return file_encoding


def get_all_file_path(in_path):
    if not os.path.exists(in_path):
        return
    path_list = os.listdir(in_path)
    for p in path_list:
        fullpath = os.path.join(in_path, p)
        if os.path.isdir(fullpath):
            get_all_file_path(fullpath)
            continue
        all_path.append(fullpath)


def init_path():
    project = os.path.abspath(os.path.dirname(os.getcwd()))
    print('项目路径:' + project)
    for p in floders:
        get_all_file_path(os.path.join(project, p))


def scan_file(in_path: str):
    if not os.path.exists(in_path):
        return
    ft = os.path.splitext(in_path)[-1]
    if ft not in files:
        return
    encode = get_file_encoding(in_path)
    if encode == 'Windows-1254':
        encode = 'utf-8'
    print('检查文件:', in_path)
    # is_changed = False
    with open(in_path, "r", encoding=encode) as f:
        if in_path.endswith('.level') or in_path.endswith('.prefab') or in_path.endswith('.mat'):
            data = {}
            try:
                data = json.load(f)
            except:
                print(in_path, '读取异常')
                return
            if 'Assets' in data:
                data['Assets'] = ''
            if 'Asset' in data:
                data['Asset'] = ''
            if 'Guid' in data:
                guid = str(data['Guid'])
            else:
                mn = str(in_path).split("\\")[-1]
                guid = mn.replace('.mat', '')
                # print(guid)
            content = json.dumps(data)
        else:
            try:
                content = f.read()
            except:
                print(in_path, '读取异常')
                return

        if in_path.endswith('.level'):
            # scan_mat(data['InstanceGroup'])
            scan_scene(data['Scene'], in_path)
        elif in_path.endswith('.prefab'):
            scan_scene(data['PrefabData']['Scene'], in_path)
        else:
            for g in DiscardedGuidArray:
                gstr = '\"' + g + '\"'
                if in_path.endswith('.asset'):
                    gstr = '\"' + g+'\\'
                idx = content.find(gstr)
                if idx != -1:
                    if in_path.endswith('.ui') and idx > 10:
                        gid = content[idx-10:idx].lower()
                        # print(gid)
                        if gid.find('guid') < 0:
                            continue
                    msg = 'path[' + in_path + '] use guid : [ ' + g + ' ]'
                    if in_path.endswith('.mat'):
                        msg = 'path[' + in_path + '] [ ' + \
                            guid+' ]use guid : [ ' + g + ' ]'
                        if guid not in mats.keys():
                            mats[guid] = in_path
                        if DelMat:
                            mat_error[guid] = msg
                        else:
                            result_error.append(msg)
                    elif in_path.endswith('.ts'):
                        if int(g) > 1100:
                            if g in replaceMap.keys():
                                vr = replaceMap[g]
                                msg += '=>[ '+vr['rid']+' ]'
                            tslist.append(msg)
                    else:
                        result_error.append(msg)

    #     for k, v in DiscardedGuidMap.items():
    #         if in_path.endswith('.level') or in_path.endswith('.prefab'):
    #             k_str = '\"Asset\":\"' + k + '\"'
    #             v_str = '\"Asset\":\"' + v + '\"'

    #             k_str_1 = '\"Asset\": \"' + k + '\"'
    #             v_str_1 = '\"Asset\": \"' + v + '\"'
    #         else:
    #             k_str = '\"' + k + '\"'
    #             v_str = '\"' + v + '\"'
    #         k_str2 = '\"' + k + ','
    #         k_str3 = ',' + k + '\"'
    #         k_str4 = ',' + k + ','

    #         if content.find(k_str) != -1:
    #             result_log.append('path[' + in_path + '] use gameplay object guid : [' + k + ']' +
    #                               ' change to [' + v + ']')
    #             content = content.replace(k_str, v_str)
    #             # is_changed = True

    #         if in_path.endswith('.level') or in_path.endswith('.prefab'):
    #             if content.find(k_str_1) != -1:
    #                 result_log.append('path[' + in_path + '] use gameplay object guid : [' + k + ']' +
    #                                   ' change to [' + v + ']')
    #                 content = content.replace(k_str_1, v_str_1)
    #                 is_changed = True

    #         if content.find(k_str2) != -1 or content.find(k_str3) != -1 or content.find(k_str4) != -1:
    #             result_warning.append('path[' + in_path + '] may use gameplay object guid : [' + k + ']' +
    #                                   ' you can change to [' + v + ']!!!')

    # if is_changed:
    #     with open(in_path, "w", encoding=encode) as f:
    #         f.write(content)


def scan_scene(scene, path):
    if not scene:
        return
    for data in scene:
        id = 'NoneGuid'
        sn = ''
        if 'Guid' in data:
            id = data['Guid']
        if 'Name' in data:
            sn = data['Name']
        msg = 'path[' + path + ']guid[ '+id + ' '+sn+' ] use '
        if 'Asset' in data:
            asset = data['Asset']
            if asset in DiscardedGuidArray:
                # print(id, asset)
                tip = msg + 'guid : [ ' + asset + ' ]'
                result_error.append(tip)
        if 'StaticMesh' in data:
            if 'Asset' in data['StaticMesh']:
                asset = data['StaticMesh']['Asset']
                if asset in DiscardedGuidArray:
                    tip = msg + 'guid : [ ' + asset + ' ]'
                    result_error.append(tip)
        if 'Script' in data:
            scriptdata = data['Script']
            if 'CharacterAsset' in scriptdata:
                info = json.dumps(scriptdata)
                for g in DiscardedGuidArray:
                    gstr = '\"' + g + '\"'
                    if info.find(gstr) != -1:
                        tip = msg + 'guid : [ ' + g + ' ]'
                        result_error.append(tip)

        for mat in mats.keys():
            info = json.dumps(data)
            if info.find(mat) != -1:
                if mat not in useMats:
                    useMats.append(mat)
                tip = msg + 'mat : [ ' + mats[mat] + ' ]'
                guids.append(tip)


if __name__ == '__main__':
    # 创建 logger 对象
    logger = logging.getLogger()
    logger.setLevel(logging.INFO)

    # 创建 console handler 并设置级别为 DEBUG
    ch = logging.StreamHandler()
    ch.setLevel(logging.INFO)

    root = os.getcwd()
    if not root.endswith('ResCheck'):
        file = root+'/ResCheck'
        os.chdir(file)
    a = time.time()
    init_path()

    if xlsx_init():
        print(' 废弃资源数量:' + str(DiscardedGuidArray.__len__()))
        print(' 检测中...')

        for path in all_path:
            scan_file(path)
            # scan_mat(path)
        # if useMats.__len__() < mats.__len__():
        #     print('————————————————————————————————————————————————————————————————————')
        #     print('        ——————————————————— 下列GUID的材质没有使用? ——————————————————')
        #     print('————————————————————————————————————————————————————————————————————')
        # for k, v in mats.items():
        #     if k not in useMats:
        #         if DelMat:
        #             print('[_DelMAT_] ['+v+'] '+k)
        #             del mat_error[k]
        #             os.remove(v)
        #         else:
        #             print('[_MAT_] ['+v+'] '+k)

        # if result_log.__len__() > 0 or result_warning.__len__() > 0:
        #     print('—————————————————————————————————————————————————————————————————————')
        #     print('        ——————————————————— 下列文件使用到废弃功能对象 ——————————————————')
        #     print('—————————————————————————————————————————————————————————————————————')
        # formatter = logging.Formatter(
        #     '[_%(levelname)s_] [ %(message)s')
        # ch.setFormatter(formatter)
        # logger.addHandler(ch)
        # for r in result_log:
        #     logger.info(r)

        # formatter = logging.Formatter(
        #     '[_%(levelname)s_] [%(message)s')
        # ch.setFormatter(formatter)
        # logger.addHandler(ch)
        # for r in result_warning:
        #     logger.warning(r)

        # formatter = logging.Formatter(
        #     '[_%(levelname)s_] [ %(message)s')
        # ch.setFormatter(formatter)
        # logger.addHandler(ch)

        ten = tslist.__len__()
        if ten > 0:
            print('————————————————————————————————————————————————————————————————————')
            print(
                '        ——————————————————— TS文件使用到废弃资源,如此ID不代表资源可以忽略——————————————————')
            print('————————————————————————————————————————————————————————————————————')
        for r in tslist:
            logger.error(r)

        en = result_error.__len__()+mat_error.__len__()
        if en > 0:
            print('————————————————————————————————————————————————————————————————————')
            print('        ——————————————————— 下列文件使用到废弃资源 ' +
                  str(en)+'——————————————————')
            print('————————————————————————————————————————————————————————————————————')
        for k, v in mat_error.items():
            logger.error(v)
        for r in result_error:
            logger.error(r)

        if guids.__len__() > 0:
            print('————————————————————————————————————————————————————————————————————')
            print('        ——————————————————— 下列GUID的对象使用了废弃资源的材质 ——————————————————')
            print('————————————————————————————————————————————————————————————————————')
        for g in guids:
            print('[_ERROR_]'+g)

        if result_log.__len__() == 0 and result_warning.__len__() == 0 and result_error.__len__() == 0:
            if ten > 0:
                print('检查上述TS文件，如果不是资源ID，则项目没有使用到废弃资源')
            else:
                print('项目没有使用到废弃资源')
        else:
            print('请修改上述文件使用到的废弃资源')
        a = time.time() - a
        print('use time : ' + str(a))
    else:
        print('文件:[' + XlsxFilePath + ']不存在')
    time.sleep(1)
    os.system('pause')
