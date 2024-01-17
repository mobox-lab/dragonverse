import json
import os
import time

import chardet
import openpyxl

# 是否替换自定义材质为默认材质
diyMat = True

resXlsx = '替换表.xlsx'
# 替换检查的目录 只替换材质、关卡、预制体
floders = ["Materials", 'Prefabs', 'Levels', 'TempPrefabs', 'Character']

replaceMap = {}
pathList = []
totle = 0
meshMap = {}


def xlsx_init():
    if not os.path.exists(resXlsx):
        return False
    wb = openpyxl.load_workbook(resXlsx, data_only=True)
    sheetName = wb.sheetnames[0]
    sheet = wb[sheetName]
    rowCount = sheet.max_row
    keys = ['oid', 'type', 'rid', 'name', 'path', 'mesh', 'mat']
    for row in range(2, rowCount + 1):
        key = sheet.cell(row, 1).value
        if key == None:
            continue
        line = {}
        col = 1
        for k in keys:
            v = sheet.cell(row, col).value
            if v and v != '0':
                line[k] = str(v)
            elif k == 'type':
                col = 0
                break
            col += 1
        if col > 0:
            replaceMap[str(key)] = line
    # print(replaceMap.keys())
    return True


def get_file_encoding(in_path):
    f = open(in_path, 'rb')
    data = f.read()
    file_encoding = chardet.detect(data).get('encoding')
    f.close()
    return file_encoding


def get_all_file_path(in_path):
    if not os.path.exists(in_path):
        return

    path_list = os.listdir(in_path)
    for p in path_list:
        fullpath = os.path.join(in_path, p)
        if os.path.isdir(fullpath):
            get_all_file_path(fullpath)
            continue
        if fullpath.endswith('.mat') or fullpath.endswith('.level') or fullpath.endswith('.prefab') or fullpath.endswith('.asset'):
            pathList.append(fullpath)


def init_path():
    project = os.path.abspath(os.path.dirname(os.getcwd()))
    print('项目路径:' + project)
    for p in floders:
        get_all_file_path(os.path.join(project, p))


def scan_file(in_path: str):
    if not os.path.exists(in_path):
        return
    encode = get_file_encoding(in_path)
    if encode == 'Windows-1254':
        encode = 'utf-8'
    global totle
    with open(in_path, "r", encoding=encode) as f:
        num = 0
        data = {}
        try:
            data = json.load(f)
            content = json.dumps(data)
        except:
            print(in_path, '读取异常')
            return

        if 'Assets' in data:
            arry = data['Assets']

        for k, v in replaceMap.items():
            rid = str(v['rid'])
            key = '\"' + k + '\"'
            res = '\"' + rid + '\"'
            count = 0
            t = v['type']
            rep = True
            meshMap.clear()
            if in_path.endswith('.level'):
                if arry and k in arry:
                    arry.remove(k)
                    data['Assets'] = arry
                    num = 1
                    print("["+in_path+"]Assets移除[ "+k+" ]")
                count = scan_scene(data['Scene'], in_path, key, res)
                if (meshMap.__len__() > 0):
                    sacn_group(data['InstanceGroup'])
                content = json.dumps(data)
                if t != '3':
                    rep = False
            elif in_path.endswith('.prefab'):
                if not 'PrefabData' in data:
                    continue
                count = scan_scene(data['PrefabData']
                                   ['Scene'], in_path, key, res)
                content = json.dumps(data)
                if t != '3':
                    rep = False
            elif in_path.endswith('.asset'):
                key = '\"' + k + '\\'
                res = '\"' + rid + '\\'
            if rep:
                count = content.count(key)
                content = content.replace(key, res)
                if count > 0:
                    print("["+in_path+"]替换[ "+key+" ]为[ "+res+" ]"+str(count)+"处")
            num += count
        totle += num
        if num > 0:
            with open(in_path, "w", encoding=encode) as f:
                f.write(content)


def scan_scene(scene, path, oid, nid):
    if not scene:
        return 0
    num = 0
    for data in scene:
        if 'Script' in data:
            scriptdata = data['Script']
            if 'CharacterAsset' in scriptdata:
                info = json.dumps(data['Script'])
                if info.find(oid) > -1:
                    num += 1
                    info = info.replace(oid, nid)
                    print("["+path+"]替换[ "+oid+" ]为[ "+nid+" ]")
                    data['Script'] = json.loads(info)
        if not 'Asset' in data:
            continue
        asset = data['Asset']
        if asset in replaceMap.keys():
            num += 1
            v = replaceMap[asset]
            t = v['type']
            if t == '0':
                continue
            rid = v['rid']
            data['Asset'] = rid
            print("["+path+"]替换[ "+asset+" ]为[ "+rid+" ]")
            if t == '4' and 'Effects' in data:
                data['Effects'] = rid
                if 'SourceType' in data:
                    if v['path']:
                        data['SourceType'] = v['path']
                    else:
                        print(rid+'没有SourceType配置')
            if t == '5' and 'SKGuid' in data:
                data['SKGuid'] = rid
            if t != '2':
                continue
            if 'SourceType' in data:
                data['SourceType'] = v['path']
            if 'StaticMesh' in data:
                sma = data['StaticMesh']['Asset']
                mas = []

                data['StaticMesh']['Asset'] = v['mesh']
                mats = data['StaticMesh']['Materials']
                if not mats:
                    continue
                bm = False
                for k, m in mats.items():
                    mat = str(m)
                    if mat.__len__() < 30 or diyMat:
                        bm = True
                if bm:
                    ms = v['mat']
                    arry = ms.split(',')
                    mv = {}
                    idx = 0
                    for k in arry:
                        mv[str(idx)] = k
                        idx += 1
                        mas.append(k)
                    # if asset == '2003':
                    #     print(arry, mv)
                    data['StaticMesh']['Materials'] = mv
                if sma not in meshMap.keys():
                    mm = {}
                    mm['mesh'] = v['mesh']
                    mm['mat'] = mas
                    meshMap[sma] = mm
    return num


def sacn_group(scene):
    if not scene:
        return
    for data in scene:
        if 'StaticMesh' in data:
            ms = str(data['StaticMesh'])
            if ms in meshMap.keys():
                mm = meshMap[ms]
                data['StaticMesh'] = mm['mesh']
                data['Materials'] = mm['mat']
                # print(ms, mm['mat'])


if __name__ == '__main__':
    a = time.time()
    root = os.getcwd()
    if not root.endswith('ResCheck'):
        file = root+'/ResCheck'
        os.chdir(file)
    init_path()
    if xlsx_init():
        print(' 替换列表:' + str(replaceMap.__len__()))
        print(' 检测中...')
        for path in pathList:
            scan_file(path)
        if totle > 0:
            print(' 废弃资源替换完成，共'+totle.__str__()+'个文件')
        else:
            print(' 没有可替换的了')
        a = time.time() - a
    else:
        print(' [替换表.xlsx]不存在')
    print(' use time : ' + str(a))
    print('-----------------------------------------------')
    print('-------------检查还有无其他废弃资源--------------')
    print('-----------------------------------------------')
    time.sleep(5)
    os.system(r'python .\ScanAsset.py')
